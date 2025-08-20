# -*- coding: utf-8 -*-
# =============================================================================
# 📦 Streamlit 식자재 발주 시스템 (v9.3 - data_editor 오류 수정)
#
# - 주요 개선사항:
#   - 장바구니 표시 로직을 안정화하여 st.data_editor에서 발생하던 오류 해결
# =============================================================================

from io import BytesIO
from datetime import datetime, date, timedelta
from typing import Dict, Any, List, Optional
from collections.abc import Mapping
from zoneinfo import ZoneInfo
import math
import hashlib
from pathlib import Path

import pandas as pd
import streamlit as st
import requests

# Google API
import gspread
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# Excel
import xlsxwriter
from openpyxl import load_workbook

# 1) 병합 셀 안전 쓰기 (이미 있으면 중복 두지 말고 그대로 사용)
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

def _safe_set(ws, row: int, col: int, value):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(mr.min_row, mr.min_col).value = value
            return
    ws.cell(row, col).value = value

# 2) 시트 안전 선택 (시트명이 바뀌거나 커버시트가 있을 때 대비)
def _pick_sheet(wb, preferred_names: list[str], fallback_contains: list[str] = []):
    # 선호 시트명 우선
    for name in preferred_names:
        if name in wb.sheetnames:
            return wb[name]
    # 제목에 특정 키워드 포함 시트
    lowers = [s.lower() for s in wb.sheetnames]
    for kw in fallback_contains:
        for i, s in enumerate(lowers):
            if kw.lower() in s:
                return wb[wb.sheetnames[i]]
    # 마지막으로 active
    return wb.active

# 3) 지점 정보 키 매핑(시트 컬럼명이 조금씩 달라도 안전하게)
def _normalize_store_info(store_info: pd.Series) -> dict:
    s = {k: ("" if pd.isna(v) else v) for k, v in store_info.to_dict().items()}
    return {
        # 사업자등록번호, 상호명, 주소, 업태(없으면 빈값)
        "사업자등록번호": s.get("사업자등록번호") or s.get("사업자번호") or s.get("등록번호") or "",
        "상호명":         s.get("상호명") or s.get("지점명") or s.get("상호") or "",
        "사업장주소":     s.get("사업장주소") or s.get("주소") or "",
        "업태":           s.get("업태") or s.get("업종") or "",
    }

# -----------------------------------------------------------------------------
# 페이지/테마/스타일
# -----------------------------------------------------------------------------
st.set_page_config(page_title="산카쿠 식자재 발주 시스템", page_icon="📦", layout="wide")

THEME = { "BORDER": "#e8e8ee", "PRIMARY": "#1C6758", "BG": "#f7f8fa", "CARD_BG": "#ffffff", "TEXT": "#222", "MUTED": "#777" }

# 탭 UI 복원을 위한 CSS (Streamlit 최신 버전 호환)
st.markdown(f"""<br><style>
    .stTabs [data-baseweb="tab-list"] {{
        gap: 12px;
    }}
    .stTabs [data-baseweb="tab"] {{
        height: 42px;
        border: 1px solid {THEME['BORDER']};
        border-radius: 12px;
        background-color: #fff;
        padding: 10px 14px;
        box-shadow: 0 1px 6px rgba(0,0,0,0.04);
    }}
    .stTabs [aria-selected="true"] {{
        border-color: {THEME['PRIMARY']};
        color: {THEME['PRIMARY']};
        box-shadow: 0 6px 16px rgba(28,103,88,0.18);
        font-weight: 700;
    }}
    html, body, [data-testid="stAppViewContainer"] {{
        background: {THEME['BG']};
        color: {THEME['TEXT']};
    }}
    .block-container {{
        padding-top: 2.4rem;
        padding-bottom: 1.6rem;
    }}
    [data-testid="stAppViewContainer"] .main .block-container {{
        max-width: 1050px;
        margin: 0 auto;
        padding: 0 12px;
    }}
    .stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] {{
        display: none;
    }}
</style>""", unsafe_allow_html=True)


def v_spacer(height: int):
    st.markdown(f"<div style='height:{height}px'></div>", unsafe_allow_html=True)

KST = ZoneInfo("Asia/Seoul")
def now_kst_str(fmt: str = "%Y-%m-%d %H:%M:%S") -> str: return datetime.now(KST).strftime(fmt)

def display_feedback():
    if "success_message" in st.session_state and st.session_state.success_message:
        st.success(st.session_state.success_message)
        st.session_state.success_message = ""

# =============================================================================
# 1) Users 로더
# =============================================================================
@st.cache_data
def load_users_from_secrets() -> Dict[str, Dict[str, Any]]:
    cleaned: Dict[str, Dict[str, str]] = {}
    users_root = st.secrets.get("users", None)
    if isinstance(users_root, Mapping) and len(users_root) > 0:
        for uid, payload in users_root.items():
            if isinstance(payload, Mapping): cleaned[str(uid)] = _normalize_account(str(uid), payload)
    if not cleaned: st.error("로그인 계정을 찾을 수 없습니다. Secrets 의 [users] 구조를 확인하세요."); st.stop()
    return cleaned

def _normalize_account(uid: str, payload: Mapping) -> dict:
    pwd_plain, pwd_hash = payload.get("password"), payload.get("password_hash")
    name = str(payload.get("name", uid)).strip()
    role = str(payload.get("role", "store")).strip().lower()
    if not (pwd_plain or pwd_hash): st.error(f"[users.{uid}]에 password 또는 password_hash가 필요합니다."); st.stop()
    if role not in {"store", "admin"}: st.error(f"[users.{uid}].role 은 'store' 또는 'admin' 이어야 합니다. (현재: {role})"); st.stop()
    return {"password": str(pwd_plain) if pwd_plain else None, "password_hash": str(pwd_hash).lower() if pwd_hash else None, "name": name, "role": role}

USERS = load_users_from_secrets()

# =============================================================================
# 2) 시트/스키마 정의
# =============================================================================
SHEET_NAME_STORES = "지점마스터"
SHEET_NAME_MASTER = "상품마스터"
SHEET_NAME_ORDERS = "발주"
SHEET_NAME_LOG = "변경로그"
MASTER_COLUMNS = ["품목코드", "품목명", "품목규격", "분류", "단위", "단가", "과세구분", "활성"]
ORDERS_COLUMNS = ["주문일시", "발주번호", "지점ID", "지점명", "납품요청일", "품목코드", "품목명", "단위", "수량", "단가", "공급가액", "세액", "합계금액", "비고", "상태", "처리일시", "처리자"]
CART_COLUMNS = ["품목코드", "품목명", "단위", "단가", "수량", "합계금액"] # 여기서 합계금액은 공급가액을 임시 저장
LOG_COLUMNS = ["변경일시", "변경자", "대상시트", "품목코드", "변경항목", "이전값", "새로운값"]

# =============================================================================
# 3) Google Sheets & Drive 연결
# =============================================================================
@st.cache_resource(show_spinner=False)
def get_google_creds():
    """Google API 인증 정보를 생성하고 반환 (Sheets와 Drive 권한 포함)"""
    google = st.secrets.get("google", {})
    creds_info = dict(google)
    if "\\n" in str(creds_info.get("private_key", "")):
        creds_info["private_key"] = str(creds_info["private_key"]).replace("\\n", "\n")
    
    # Google Sheets와 Google Drive 읽기 권한을 함께 요청
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.readonly"
    ]
    creds = service_account.Credentials.from_service_account_info(creds_info, scopes=scopes)
    return creds

@st.cache_resource(show_spinner=False)
def get_gs_client():
    """gspread 클라이언트 반환"""
    creds = get_google_creds()
    return gspread.authorize(creds)

@st.cache_resource(show_spinner="API 서비스 연결 중...")
def get_drive_service():
    """Google Drive API 서비스 클라이언트 반환"""
    creds = get_google_creds()
    return build('drive', 'v3', credentials=creds)

def download_template_from_drive(file_id: str) -> Optional[BytesIO]:
    """Google Drive에서 파일 ID를 이용해 템플릿 파일을 다운로드"""
    try:
        service = get_drive_service()
        request = service.files().get_media(fileId=file_id)
        file_buffer = BytesIO()
        downloader = MediaIoBaseDownload(file_buffer, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        file_buffer.seek(0)
        return file_buffer
    except Exception as e:
        st.error(f"Google Drive에서 템플릿 파일(ID: {file_id}) 다운로드에 실패했습니다.")
        st.error(e)
        return None

@st.cache_resource(show_spinner=False)
def open_spreadsheet():
    key = str(st.secrets.get("google", {}).get("SPREADSHEET_KEY", "")).strip()
    if not key: st.error("Secrets 에 SPREADSHEET_KEY가 없습니다."); st.stop()
    try: return get_gs_client().open_by_key(key)
    except Exception as e: st.error(f"스프레드시트 열기 실패: {e}"); st.stop()

# =============================================================================
# 4) 데이터 I/O 함수
# =============================================================================
@st.cache_data(ttl=3600)
def load_store_info_df() -> pd.DataFrame:
    try:
        ws = open_spreadsheet().worksheet(SHEET_NAME_STORES)
        df = pd.DataFrame(ws.get_all_records(empty2zero=False))
        return df
    except gspread.WorksheetNotFound:
        st.error(f"'{SHEET_NAME_STORES}' 시트를 찾을 수 없습니다."); return pd.DataFrame()

@st.cache_data(ttl=180)
def load_master_df() -> pd.DataFrame:
    try:
        ws = open_spreadsheet().worksheet(SHEET_NAME_MASTER)
        df = pd.DataFrame(ws.get_all_records(empty2zero=False))
        mask = df["활성"].astype(str).str.lower().isin(["1", "true", "y", "yes", ""])
        df = df[mask | df["활성"].isna()]
        df["단가"] = pd.to_numeric(df["단가"], errors="coerce").fillna(0).astype(int)
        return df
    except gspread.WorksheetNotFound:
        st.error(f"'{SHEET_NAME_MASTER}' 시트를 찾을 수 없습니다."); return pd.DataFrame()

def write_master_df(df: pd.DataFrame, original_df: pd.DataFrame) -> bool:
    st.error("write_master_df 함수가 구현되지 않았습니다.")
    return False

@st.cache_data(ttl=60)
def load_orders_df() -> pd.DataFrame:
    try:
        ws = open_spreadsheet().worksheet(SHEET_NAME_ORDERS)
        df = pd.DataFrame(ws.get_all_records(empty2zero=False))
        for c in ORDERS_COLUMNS:
            if c not in df.columns: df[c] = ""
        money_cols = ["수량", "단가", "공급가액", "세액", "합계금액"]
        for c in money_cols: df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
        df = df.sort_values(by="주문일시", ascending=False)
        return df[ORDERS_COLUMNS].copy()
    except gspread.WorksheetNotFound:
        st.warning(f"'{SHEET_NAME_ORDERS}' 시트가 없어 새로 생성합니다. 첫 발주 후 확인해주세요.")
        return pd.DataFrame(columns=ORDERS_COLUMNS)

# [수정] append_orders 함수 구현
def append_orders(rows: List[Dict[str, Any]]) -> bool:
    """지정된 행(List[Dict])을 '발주' 시트에 추가합니다."""
    if not rows:
        return True
    try:
        ss = open_spreadsheet()
        ws = ss.worksheet(SHEET_NAME_ORDERS)
        
        values_to_append = []
        for row_dict in rows:
            row_list = [row_dict.get(col, "") for col in ORDERS_COLUMNS]
            values_to_append.append(row_list)
            
        ws.append_rows(values_to_append, value_input_option='USER_ENTERED')
        st.cache_data.clear()
        return True
    except gspread.exceptions.WorksheetNotFound:
        st.error(f"'{SHEET_NAME_ORDERS}' 시트를 찾을 수 없습니다.")
        return False
    except Exception as e:
        st.error(f"발주 데이터 추가 중 오류가 발생했습니다: {e}")
        return False

def append_change_log(log_entries: List[Dict[str, Any]]):
    st.error("append_change_log 함수가 구현되지 않았습니다.")
    return True

# [수정] update_order_status 함수 구현
def update_order_status(selected_ids: List[str], new_status: str, handler: str) -> bool:
    """선택된 발주번호들의 상태를 일괄 변경합니다."""
    if not selected_ids:
        return True
    try:
        ss = open_spreadsheet()
        ws = ss.worksheet(SHEET_NAME_ORDERS)
        
        all_data = ws.get_all_values()
        header = all_data[0]
        records = all_data[1:]

        id_col_idx = header.index("발주번호") + 1
        status_col_idx = header.index("상태") + 1
        handler_col_idx = header.index("처리자") + 1
        timestamp_col_idx = header.index("처리일시") + 1
        
        cells_to_update = []
        now_str = now_kst_str()
        
        for i, row in enumerate(records, start=2):
            if row[id_col_idx - 1] in selected_ids:
                cells_to_update.append(gspread.Cell(i, status_col_idx, new_status))
                cells_to_update.append(gspread.Cell(i, handler_col_idx, handler))
                cells_to_update.append(gspread.Cell(i, timestamp_col_idx, now_str))

        if cells_to_update:
            ws.update_cells(cells_to_update, value_input_option='USER_ENTERED')
        
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"발주 상태 업데이트 중 오류가 발생했습니다: {e}")
        return False

# =============================================================================
# 5) 로그인
# =============================================================================
def require_login():
    if st.session_state.get("auth", {}).get("login"): return True
    st.markdown('<div style="text-align:center; font-size:42px; font-weight:800; margin:16px 0 12px;">식자재 발주 시스템</div>', unsafe_allow_html=True)
    _, mid, _ = st.columns([3, 2, 3])
    with mid.form("login_form"):
        uid = st.text_input("아이디 또는 지점명", key="login_uid", placeholder="예: jeondae / 전대점")
        pwd = st.text_input("비밀번호", type="password", key="login_pw")
        if st.form_submit_button("로그인", use_container_width=True):
            real_uid, acct = _find_account(uid)
            if not (real_uid and acct and verify_password(pwd, acct.get("password_hash"), acct.get("password"))):
                st.error("아이디(또는 지점명) 또는 비밀번호가 올바르지 않습니다.")
            else:
                st.session_state["auth"] = {"login": True, "user_id": real_uid, "name": acct["name"], "role": acct["role"]}; st.rerun()
    return False

def verify_password(input_pw: str, stored_hash: Optional[str], fallback_plain: Optional[str]) -> bool:
    if stored_hash: return hashlib.sha256(input_pw.encode()).hexdigest() == stored_hash.strip().lower()
    return str(input_pw) == str(fallback_plain) if fallback_plain is not None else False

def _find_account(uid_or_name: str):
    s_lower = str(uid_or_name or "").strip().lower()
    if not s_lower: return None, None
    for uid, acct in USERS.items():
        if uid.lower() == s_lower or acct.get("name", "").lower() == s_lower: return uid, acct
    return None, None
    
# =============================================================================
# 6) 템플릿 기반 Excel 생성 함수들
# =============================================================================
def make_order_id(store_id: str) -> str: return f"{datetime.now(KST):%Y%m%d%H%M%S}{store_id}"

def make_trading_statement_excel(df_doc: pd.DataFrame, store_info: pd.Series, master_df: pd.DataFrame) -> BytesIO:
    if df_doc.empty:
        st.warning("거래명세서를 생성할 데이터가 없습니다.")
        return BytesIO()
        
    total_amount = int(pd.to_numeric(df_doc["합계금액"], errors="coerce").fillna(0).sum())

    base_dt = pd.to_datetime(df_doc["납품요청일"].iloc[0], errors="coerce")
    if pd.isna(base_dt): base_dt = pd.Timestamp.now(tz=KST)

    template_id = st.secrets.get("google", {}).get("TEMPLATE_TRADING_STATEMENT_ID")
    if not template_id:
        st.error("Secrets에 거래명세서 템플릿 ID(TEMPLATE_TRADING_STATEMENT_ID)가 없습니다.")
        return BytesIO()
    template_bytes = download_template_from_drive(template_id)
    if template_bytes is None: return BytesIO()

    wb = load_workbook(template_bytes, data_only=False)
    ws = _pick_sheet(wb, ["거래명세서"], ["명세", "trading", "statement"])

    supplier = { "등록번호": "686-85-02906", "상호": "산카쿠 대전 가공장", "성명": "이수정", "사업장": "대전광역시 서구 둔산로18번길 62, 101호", }
    store_norm = _normalize_store_info(store_info)

    _safe_set(ws, 4, 18, supplier["등록번호"])
    _safe_set(ws, 6, 18, supplier["상호"])
    _safe_set(ws, 6, 22, supplier["성명"])
    _safe_set(ws, 8, 18, supplier["사업장"])
    _safe_set(ws, 4,  3, store_norm["상호명"])
    _safe_set(ws, 6,  3, store_norm["사업장주소"])
    _safe_set(ws,10, 3, int(total_amount))

    COL_YEAR, COL_MONTH, COL_DAY = 2, 3, 4
    COL_ITEM, COL_SPEC = 5, 13
    COL_QTY,  COL_UNIT = 18, 22
    COL_SUP,  COL_TAX = 26, 31

    start_row = 13
    df_m = pd.merge(df_doc, master_df[["품목코드", "품목규격"]], on="품목코드", how="left")

    r = start_row
    for _, row in df_m.iterrows():
        d = pd.to_datetime(row.get("납품요청일"), errors="coerce")
        if pd.isna(d): d = base_dt

        _safe_set(ws, r, COL_YEAR,  int(d.year))
        _safe_set(ws, r, COL_MONTH, int(d.month))
        _safe_set(ws, r, COL_DAY,   int(d.day))
        _safe_set(ws, r, COL_ITEM,  str(row["품목명"]))
        _safe_set(ws, r, COL_SPEC,  str(row.get("품목규격", "") or ""))
        _safe_set(ws, r, COL_QTY,   int(pd.to_numeric(row["수량"], errors="coerce") or 0))
        _safe_set(ws, r, COL_UNIT,  int(pd.to_numeric(row["단가"], errors="coerce") or 0))

        ws.cell(r, COL_SUP).value = f'=IF({get_column_letter(COL_UNIT)}{r}="","",{get_column_letter(COL_QTY)}{r}*{get_column_letter(COL_UNIT)}{r})'
        ws.cell(r, COL_TAX).value = f'=IF({get_column_letter(COL_UNIT)}{r}="","",{get_column_letter(COL_SUP)}{r}*0.1)'
        r += 1

    ws.print_area = "$B$2:$AH$53"
    ps = ws.page_setup
    ps.paperSize, ps.orientation, ps.fitToWidth, ps.fitToHeight, ps.scale = 9, "portrait", 1, 0, None
    ws.page_margins.left, ws.page_margins.right, ws.page_margins.top, ws.page_margins.bottom = 0.25, 0.25, 0.5, 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out = BytesIO(); wb.save(out); out.seek(0)
    return out

def make_tax_invoice_excel(df_doc: pd.DataFrame, store_info: pd.Series, master_df: pd.DataFrame) -> BytesIO:
    if df_doc.empty:
        st.warning("세금계산서를 생성할 데이터가 없습니다.")
        return BytesIO()

    try: base_dt = pd.to_datetime(df_doc["납품요청일"].iloc[0], errors="coerce")
    except Exception: base_dt = None
    if pd.isna(base_dt): base_dt = pd.Timestamp.now(tz=KST)

    template_id = st.secrets.get("google", {}).get("TEMPLATE_TAX_INVOICE_ID")
    if not template_id:
        st.error("Secrets에 세금계산서 템플릿 ID(TEMPLATE_TAX_INVOICE_ID)가 없습니다.")
        return BytesIO()
    template_bytes = download_template_from_drive(template_id)
    if template_bytes is None: return BytesIO()

    wb = load_workbook(template_bytes, data_only=False)
    ws = _pick_sheet(wb, ["세금계산서양식", "세금계산서"], ["tax", "invoice", "계산서"])
    store_norm = _normalize_store_info(store_info)

    supplier = { "등록번호": "686-85-02906", "상호": "산카쿠 대전 가공장", "사업장": "대전광역시 서구 둔산로18번길 62, 101호", "업태": "제조업" }
    _safe_set(ws,  5,  6, supplier["등록번호"])
    _safe_set(ws,  7,  6, supplier["상호"])
    _safe_set(ws,  9,  6, supplier["사업장"])
    _safe_set(ws, 11,  6, supplier["업태"])
    _safe_set(ws,  5, 19, store_norm["사업자등록번호"])
    _safe_set(ws,  8, 19, store_norm["상호명"])
    _safe_set(ws, 10, 19, store_norm["사업장주소"])
    _safe_set(ws, 11, 19, store_norm["업태"])

    for base_row in (14, 40):
        _safe_set(ws, base_row, 2, int(base_dt.year))
        _safe_set(ws, base_row, 4, int(base_dt.month))
        _safe_set(ws, base_row, 5, int(base_dt.day))

    items_count = len(df_doc)
    first_name = str(df_doc.iloc[0]["품목명"]) if items_count else ""
    summary = (f"{first_name} 등 {items_count}건") if items_count >= 2 else first_name
    _safe_set(ws, 18, 4, summary)
    _safe_set(ws, 42, 4, summary)

    ws.print_area = "$A$4:$V$48"
    ps = ws.page_setup
    ps.paperSize, ps.orientation, ps.fitToWidth, ps.fitToHeight, ps.scale = 9, "portrait", 1, 0, None
    ws.page_margins.left, ws.page_margins.right, ws.page_margins.top, ws.page_margins.bottom = 0.25, 0.25, 0.5, 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    out = BytesIO(); wb.save(out); out.seek(0)
    return out

def make_sales_summary_excel(daily_pivot: pd.DataFrame, monthly_pivot: pd.DataFrame, title: str) -> BytesIO:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as writer:
        daily_pivot.to_excel(writer, sheet_name='일별매출현황')
        monthly_pivot.to_excel(writer, sheet_name='월별매출현황')
        
        workbook = writer.book
        h_format = workbook.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        header_format = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'align': 'center'})
        money_format = workbook.add_format({'num_format': '#,##0', 'border': 1})
        
        for name, pivot_df in [('일별매출현황', daily_pivot), ('월별매출현황', monthly_pivot)]:
            worksheet = writer.sheets[name]
            worksheet.set_zoom(90)
            worksheet.merge_range(0, 0, 0, len(pivot_df.columns), f"거래처별 {name}", h_format)
            for col_num, value in enumerate(pivot_df.columns.values):
                worksheet.write(2, col_num + 1, value, header_format)
            worksheet.write(2, 0, pivot_df.index.name, header_format)
            worksheet.set_column(0, len(pivot_df.columns), 14)
            worksheet.conditional_format(3, 1, len(pivot_df) + 2, len(pivot_df.columns), 
                                         {'type': 'no_blanks', 'format': money_format})
    return buf

# =============================================================================
# 7) 장바구니 유틸
# =============================================================================
def init_session_state():
    defaults = {"cart": pd.DataFrame(columns=CART_COLUMNS), "store_editor_ver": 0, "success_message": ""}
    for key, value in defaults.items():
        if key not in st.session_state: st.session_state[key] = value

def coerce_cart_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in CART_COLUMNS:
        if col not in out.columns: out[col] = 0 if col in ["단가", "수량", "합계금액"] else ""
    out["수량"] = pd.to_numeric(out["수량"], errors="coerce").fillna(0).astype(int)
    out["단가"] = pd.to_numeric(out["단가"], errors="coerce").fillna(0).astype(int)
    out["합계금액"] = out["단가"] * out["수량"] # 공급가액 계산
    return out

def add_to_cart(rows_df: pd.DataFrame):
    add = rows_df[rows_df["수량"] > 0].copy()
    if add.empty: return
    add["합계금액"] = add["단가"] * add["수량"] # 공급가액 계산
    cart = st.session_state.cart.copy()
    merged = pd.concat([cart, add]).groupby("품목코드", as_index=False).agg({"품목명": "last", "단위": "last", "단가": "last", "수량": "sum"})
    merged["합계금액"] = merged["단가"] * merged["수량"] # 공급가액 계산
    st.session_state.cart = merged[CART_COLUMNS]

# =============================================================================
# 8) 지점(Store) 페이지
# =============================================================================
def page_store_register_confirm(master_df: pd.DataFrame):
    st.subheader("🛒 발주 요청")
    v_spacer(10)
    with st.container(border=True):
        st.markdown("##### 🗓️ 납품 요청 정보")
        today = date.today(); c1, c2 = st.columns([1, 1.2])
        quick = c1.radio("납품 선택", ["오늘", "내일", "직접선택"], horizontal=True, label_visibility="collapsed", key="store_reg_quick_radio")
        if quick == "오늘": 납품요청일 = today
        elif quick == "내일": 납품요청일 = today + timedelta(days=1)
        else: 납품요청일 = c2.date_input("납품 요청일", value=today, min_value=today, max_value=today + timedelta(days=7), label_visibility="collapsed", key="store_reg_date_input")
        memo = st.text_area("요청 사항(선택)", height=80, placeholder="예) 입고 시 얼음팩 추가 부탁드립니다.", key="store_reg_memo")
    v_spacer(16)
    with st.container(border=True):
        st.markdown("##### 🧾 발주 수량 입력")
        l, r = st.columns([2, 1])
        keyword = l.text_input("품목 검색(이름/코드)", placeholder="오이, P001 등", key="store_reg_keyword")
        cat_opt = ["(전체)"] + sorted(master_df["분류"].dropna().unique().tolist())
        cat_sel = r.selectbox("분류(선택)", cat_opt, key="store_reg_category")
        df_view = master_df.copy()
        if keyword: df_view = df_view[df_view.apply(lambda row: keyword.strip().lower() in str(row["품목명"]).lower() or keyword.strip().lower() in str(row["품목코드"]).lower(), axis=1)]
        if cat_sel != "(전체)": df_view = df_view[df_view["분류"] == cat_sel]
        with st.form(key="add_to_cart_form"):
            df_edit = df_view[["품목코드", "품목명", "단위", "단가", "과세구분"]].copy()
            
            df_edit["단가(VAT포함)"] = df_edit.apply(lambda row: row['단가'] * 1.1 if row['과세구분'] == '과세' else row['단가'], axis=1).astype(int)
            df_edit["수량"] = 0
            
            df_edit.rename(columns={"단가": "단가(원)"}, inplace=True)
            
            edited_disp = st.data_editor(
                df_edit[["품목코드", "품목명", "단위", "단가(원)", "단가(VAT포함)", "수량"]], 
                key=f"editor_v{st.session_state.store_editor_ver}", 
                hide_index=True, 
                disabled=["품목코드", "품목명", "단위", "단가(원)", "단가(VAT포함)"], 
                use_container_width=True, 
                column_config={"단가(원)": st.column_config.NumberColumn(), "단가(VAT포함)": st.column_config.NumberColumn(), "수량": st.column_config.NumberColumn(min_value=0)}
            )
            if st.form_submit_button("장바구니 추가", use_container_width=True, type="primary"):
                edited_disp.rename(columns={"단가(원)": "단가"}, inplace=True)
                items_to_add = coerce_cart_df(edited_disp)
                if not items_to_add[items_to_add["수량"] > 0].empty:
                    add_to_cart(items_to_add); st.session_state.store_editor_ver += 1
                st.rerun()
    v_spacer(16)
    with st.container(border=True):
        st.markdown("##### 🧺 장바구니")
        cart = st.session_state.cart
        if not cart.empty:
            # [수정] data_editor 오류 해결을 위해 데이터프레임 생성 로직 간소화
            cart_display = pd.merge(cart, master_df[['품목코드', '과세구분']], on='품목코드', how='left')
            # '합계금액' 컬럼(공급가액)을 '공급가액'으로 이름 변경
            cart_display.rename(columns={"합계금액": "공급가액"}, inplace=True)
            
            cart_display['합계금액(VAT포함)'] = cart_display.apply(
                lambda row: row['공급가액'] + math.ceil(row['공급가액'] * 0.1) if row.get('과세구분') == '과세' else row['공급가액'],
                axis=1
            ).astype(int)
            
            cart_display.rename(columns={"단가": "단가(원)", "공급가액": "공급가액(원)"}, inplace=True)

            edited_cart = st.data_editor(
                cart_display[["품목코드", "품목명", "단위", "단가(원)", "수량", "공급가액(원)", "합계금액(VAT포함)"]],
                key="cart_editor", hide_index=True, 
                disabled=["품목코드", "품목명", "단위", "단가(원)", "공급가액(원)", "합계금액(VAT포함)"],
                column_config={
                    "단가(원)": st.column_config.NumberColumn(), "수량": st.column_config.NumberColumn(min_value=0), 
                    "공급가액(원)": st.column_config.NumberColumn(), "합계금액(VAT포함)": st.column_config.NumberColumn()
                }
            )
            # 세션 상태 저장을 위해 컬럼명 원복
            edited_cart.rename(columns={"단가(원)": "단가", "공급가액(원)": "합계금액"}, inplace=True)
            st.session_state.cart = coerce_cart_df(edited_cart)
            if st.button("장바구니 비우기", use_container_width=True): st.session_state.cart = pd.DataFrame(columns=CART_COLUMNS); st.rerun()
        else: st.info("장바구니가 비어 있습니다.")
    v_spacer(16)
    with st.form("submit_form"):
        cart_now = st.session_state.cart
        
        cart_with_master = pd.merge(cart_now, master_df[['품목코드', '과세구분']], on='품목코드', how='left')
        cart_with_master['공급가액'] = cart_with_master['단가'] * cart_with_master['수량']
        cart_with_master['최종합계'] = cart_with_master.apply(
            lambda row: row['공급가액'] + math.ceil(row['공급가액'] * 0.1) if row['과세구분'] == '과세' else row['공급가액'], axis=1
        )
        total_final_amount_sum = cart_with_master['최종합계'].sum()

        st.markdown(f"**최종 확인:** 총 {len(cart_now)}개 품목, 최종 합계금액(VAT포함) **{total_final_amount_sum:,.0f}원**")
        confirm = st.checkbox("위 내용으로 발주를 제출합니다.")
        if st.form_submit_button("📦 발주 제출", type="primary", use_container_width=True, disabled=cart_now.empty):
            if not confirm: st.warning("제출 확인 체크박스를 선택해주세요."); st.stop()
            user = st.session_state.auth; order_id = make_order_id(user["user_id"])
            rows = []
            for _, r in cart_with_master.iterrows():
                unit_price = r['단가']
                quantity = r['수량']
                tax_type = r.get('과세구분', '과세')
                
                supply_price = unit_price * quantity
                tax = math.ceil(supply_price * 0.1) if tax_type == '과세' else 0
                total_amount = supply_price + tax

                rows.append({
                    "주문일시": now_kst_str(), "발주번호": order_id, "지점ID": user["user_id"], "지점명": user["name"], 
                    "납품요청일": f"{납품요청일:%Y-%m-%d}", "품목코드": r["품목코드"], "품목명": r["품목명"], 
                    "단위": r["단위"], "수량": quantity, "단가": unit_price, "공급가액": supply_price, "세액": tax, 
                    "합계금액": total_amount, "비고": memo, "상태": "접수"
                })

            if append_orders(rows):
                st.session_state.success_message = "발주가 성공적으로 제출되었습니다."; st.session_state.cart = pd.DataFrame(columns=CART_COLUMNS); st.rerun()
            else: st.error("발주 제출 중 오류가 발생했습니다.")

def page_store_orders_change(store_info_df: pd.DataFrame, master_df: pd.DataFrame):
    st.subheader("🧾 발주 조회·수정")
    display_feedback()
    df_all, user = load_orders_df(), st.session_state.auth
    df_user = df_all[df_all["지점ID"] == user["user_id"]]
    if df_user.empty: st.info("발주 데이터가 없습니다."); return
    
    c1, c2, c3 = st.columns(3)
    dt_from = c1.date_input("조회 시작일", date.today() - timedelta(days=30), key="store_orders_from")
    dt_to = c2.date_input("조회 종료일", date.today(), key="store_orders_to")
    order_id_search = c3.text_input("발주번호로 검색", key="store_orders_search", placeholder="전체 또는 일부 입력")

    df_filtered = df_user.copy()
    if order_id_search:
        df_filtered = df_filtered[df_filtered["발주번호"].str.contains(order_id_search, na=False)]
    else:
        df_filtered = df_filtered[(pd.to_datetime(df_filtered["납품요청일"]).dt.date >= dt_from) & (pd.to_datetime(df_filtered["납품요청일"]).dt.date <= dt_to)]

    orders = df_filtered.groupby("발주번호").agg(주문일시=("주문일시", "first"), 건수=("품목코드", "count"), 합계금액=("합계금액", "sum"), 상태=("상태", "first")).reset_index().sort_values("주문일시", ascending=False)
    orders.rename(columns={"합계금액": "합계금액(원)"}, inplace=True)
    pending = orders[orders["상태"] == "접수"].copy(); shipped = orders[orders["상태"] == "출고완료"].copy()
    
    if 'store_pending_selection' not in st.session_state: st.session_state.store_pending_selection = {}
    if 'store_shipped_selection' not in st.session_state: st.session_state.store_shipped_selection = {}

    tab1, tab2 = st.tabs([f"접수 ({len(pending)}건)", f"출고완료 ({len(shipped)}건)"])
    
    with tab1:
        if st.session_state.get('active_store_tab') != 'pending':
            st.session_state.store_shipped_selection = {}
        st.session_state.active_store_tab = 'pending'
            
        pending.insert(0, "선택", pending['발주번호'].apply(lambda x: st.session_state.store_pending_selection.get(x, False)))
        edited_pending = st.data_editor(pending, key="store_pending_editor", hide_index=True, disabled=["발주번호", "주문일시", "건수", "합계금액(원)", "상태"], column_order=("선택", "발주번호", "주문일시", "건수", "합계금액(원)", "상태"), column_config={"합계금액(원)": st.column_config.NumberColumn(), "선택": st.column_config.CheckboxColumn(width="small")})
        st.session_state.store_pending_selection = dict(zip(edited_pending['발주번호'], edited_pending['선택']))
        selected_pending_ids = [k for k, v in st.session_state.store_pending_selection.items() if v]

        if st.button("선택 발주 삭제", disabled=not selected_pending_ids, key="delete_pending_btn"):
            if update_order_status(selected_pending_ids, "삭제", user["name"]):
                st.session_state.success_message = f"{len(selected_pending_ids)}건의 발주가 삭제되었습니다."; st.rerun()
    
    with tab2:
        if st.session_state.get('active_store_tab') != 'shipped':
            st.session_state.store_pending_selection = {}
        st.session_state.active_store_tab = 'shipped'
        
        shipped.insert(0, "선택", shipped['발주번호'].apply(lambda x: st.session_state.store_shipped_selection.get(x, False)))
        edited_shipped = st.data_editor(shipped, key="store_shipped_editor", hide_index=True, disabled=["발주번호", "주문일시", "건수", "합계금액(원)", "상태"], column_order=("선택", "발주번호", "주문일시", "건수", "합계금액(원)", "상태"), column_config={"합계금액(원)": st.column_config.NumberColumn(), "선택": st.column_config.CheckboxColumn(width="small")})
        st.session_state.store_shipped_selection = dict(zip(edited_shipped['발주번호'], edited_shipped['선택']))
        selected_shipped_ids = [k for k, v in st.session_state.store_shipped_selection.items() if v]

    v_spacer(16)
    with st.container(border=True):
        st.markdown("##### 📄 발주 품목 상세 조회")
        total_selected = selected_pending_ids + selected_shipped_ids
        if len(total_selected) == 1:
            target_id = total_selected[0]
            target_df = df_user[df_user["발주번호"] == target_id]
            target_status = target_df.iloc[0]["상태"]
            
            df_display = target_df.copy().rename(columns={"단가": "단가(원)", "공급가액": "공급가액(원)", "세액": "세액(원)", "합계금액": "합계금액(원)"})
            display_cols = ["품목코드", "품목명", "단위", "수량", "단가(원)", "공급가액(원)", "세액(원)", "합계금액(원)"]
            
            st.dataframe(df_display[display_cols], hide_index=True, use_container_width=True, 
                         column_config={
                             "단가(원)": st.column_config.NumberColumn(), "공급가액(원)": st.column_config.NumberColumn(), 
                             "세액(원)": st.column_config.NumberColumn(), "합계금액(원)": st.column_config.NumberColumn()
                         })
            
            if target_status == '출고완료':
                v_spacer(10)
                store_info_series = store_info_df[store_info_df["지점ID"] == user["user_id"]]
                if not store_info_series.empty:
                    store_info = store_info_series.iloc[0]
                    buf = make_trading_statement_excel(target_df, store_info, master_df)
                    st.download_button(f"'{target_id}' 거래명세서 다운로드", data=buf, file_name=f"거래명세서_{user['name']}_{target_id}.xlsx", mime="application/vnd.ms-excel", use_container_width=True)
        else:
            st.info("상세 내용을 보려면 위 목록에서 발주를 **하나만** 선택하세요.")

def page_store_documents(store_info_df: pd.DataFrame, master_df: pd.DataFrame):
    st.subheader("📑 증빙서류 다운로드")
    user = st.session_state.auth
    df = load_orders_df()
    df_completed = df[(df["지점ID"] == user["user_id"]) & (df["상태"] == "출고완료")]
    if df_completed.empty: st.info("'출고완료' 상태의 발주 데이터가 없습니다."); return

    search_mode = st.radio("조회 방식", ["기간으로 조회", "발주번호로 조회"], key="store_doc_search_mode", horizontal=True)
    dfv = pd.DataFrame(); doc_type = "거래명세서"
    if search_mode == "기간으로 조회":
        c1, c2, c3 = st.columns([1, 1, 2])
        dt_from = c1.date_input("조회 시작일", date.today() - timedelta(days=30), key="store_doc_from")
        dt_to = c2.date_input("조회 종료일", date.today(), key="store_doc_to")
        doc_type = c3.selectbox("문서 종류", ["거래명세서", "세금계산서"], key="store_doc_type")
        mask = (pd.to_datetime(df_completed["납품요청일"]).dt.date >= dt_from) & (pd.to_datetime(df_completed["납품요청일"]).dt.date <= dt_to)
        dfv = df_completed[mask].copy()
    else:
        c1, c2 = st.columns([1, 1])
        order_ids = sorted(df_completed["발주번호"].dropna().unique().tolist(), reverse=True)
        order_id_sel = c1.selectbox("발주번호 선택", [""] + order_ids, key="store_doc_order_id")
        doc_type = c2.selectbox("문서 종류", ["거래명세서", "세금계산서"], key="store_doc_type_by_id")
        if order_id_sel: dfv = df_completed[df_completed["발주번호"] == order_id_sel].copy()

    if dfv.empty: st.warning("해당 조건으로 조회된 데이터가 없습니다."); return
    st.dataframe(dfv, use_container_width=True, hide_index=True)
    
    if not dfv.empty:
        store_info_series = store_info_df[store_info_df["지점ID"] == user["user_id"]]
        if not store_info_series.empty:
            store_info = store_info_series.iloc[0]
            if doc_type == "거래명세서":
                buf = make_trading_statement_excel(dfv, store_info, master_df)
            else:
                buf = make_tax_invoice_excel(dfv, store_info, master_df)
            st.download_button(f"{doc_type} 다운로드", data=buf, file_name=f"{doc_type}_{user['name']}_{now_kst_str('%Y%m%d')}.xlsx", mime="application/vnd.ms-excel", use_container_width=True, type="primary")
        else: 
            st.error(f"'{SHEET_NAME_STORES}' 시트에서 현재 로그인된 지점 ID '{user['user_id']}'와 일치하는 데이터를 찾을 수 없습니다. '지점ID'를 확인해주세요.")

def page_store_master_view(master_df: pd.DataFrame):
    st.subheader("🏷️ 품목 단가 조회")
    master_df_display = master_df.copy()
    # [수정] VAT 포함 단가 컬럼 추가
    master_df_display['단가(VAT포함)'] = master_df_display.apply(
        lambda row: row['단가'] * 1.1 if row['과세구분'] == '과세' else row['단가'], axis=1
    ).astype(int)
    master_df_display = master_df_display.rename(columns={"단가": "단가(원)"})
    st.dataframe(master_df_display[["품목코드", "품목명", "품목규격", "분류", "단위", "단가(원)", "단가(VAT포함)"]], use_container_width=True, hide_index=True, column_config={"단가(원)": st.column_config.NumberColumn(), "단가(VAT포함)": st.column_config.NumberColumn()})

# =============================================================================
# 9) 관리자(Admin) 페이지
# =============================================================================

def page_admin_unified_management(df_all: pd.DataFrame, store_info_df: pd.DataFrame, master_df: pd.DataFrame):
    st.subheader("📋 발주요청 조회·수정")
    display_feedback()
    if df_all.empty: st.info("발주 데이터가 없습니다."); return

    c1, c2, c3, c4 = st.columns(4)
    dt_from = c1.date_input("시작일", date.today() - timedelta(days=7), key="admin_mng_from")
    dt_to = c2.date_input("종료일", date.today(), key="admin_mng_to")
    stores = ["(전체)"] + sorted(df_all["지점명"].dropna().unique().tolist())
    store = c3.selectbox("지점", stores, key="admin_mng_store")
    order_id_search = c4.text_input("발주번호로 검색", key="admin_mng_order_id", placeholder="전체 또는 일부 입력")

    df = df_all.copy()
    if order_id_search:
        df = df[df["발주번호"].str.contains(order_id_search, na=False)]
    else:
        df = df[(pd.to_datetime(df["납품요청일"]).dt.date >= dt_from) & (pd.to_datetime(df["납품요청일"]).dt.date <= dt_to)]
        if store != "(전체)": df = df[df["지점명"] == store]

    orders = df.groupby("발주번호").agg(주문일시=("주문일시", "first"), 지점명=("지점명", "first"), 건수=("품목코드", "count"), 합계금액=("합계금액", "sum"), 상태=("상태", "first")).reset_index().sort_values("주문일시", ascending=False)
    orders.rename(columns={"합계금액": "합계금액(원)"}, inplace=True)
    pending = orders[orders["상태"] == "접수"].copy(); shipped = orders[orders["상태"] == "출고완료"].copy()
    
    if 'admin_pending_selection' not in st.session_state: st.session_state.admin_pending_selection = {}
    if 'admin_shipped_selection' not in st.session_state: st.session_state.admin_shipped_selection = {}
    
    tab1, tab2 = st.tabs([f"📦 발주 요청 접수 ({len(pending)}건)", f"✅ 출고 완료 ({len(shipped)}건)"])
    with tab1:
        if st.session_state.get('active_admin_tab') != 'pending':
            st.session_state.admin_shipped_selection = {}
        st.session_state.active_admin_tab = 'pending'

        pending.insert(0, '선택', pending['발주번호'].apply(lambda x: st.session_state.admin_pending_selection.get(x, False)))
        edited_pending = st.data_editor(pending, key="admin_pending_editor", hide_index=True, disabled=pending.columns.drop("선택"), column_order=("선택", "발주번호", "주문일시", "지점명", "건수", "합계금액(원)", "상태"), column_config={"합계금액(원)": st.column_config.NumberColumn()})
        st.session_state.admin_pending_selection = dict(zip(edited_pending['발주번호'], edited_pending['선택']))
        selected_pending_ids = [k for k, v in st.session_state.admin_pending_selection.items() if v]

        if st.button("✅ 선택 발주 출고", disabled=not selected_pending_ids, key="admin_ship_btn"):
            if update_order_status(selected_pending_ids, "출고완료", st.session_state.auth["name"]):
                st.session_state.success_message = f"{len(selected_pending_ids)}건이 출고 처리되었습니다."; st.rerun()
    with tab2:
        if st.session_state.get('active_admin_tab') != 'shipped':
            st.session_state.admin_pending_selection = {}
        st.session_state.active_admin_tab = 'shipped'

        shipped.insert(0, '선택', shipped['발주번호'].apply(lambda x: st.session_state.admin_shipped_selection.get(x, False)))
        edited_shipped = st.data_editor(shipped, key="admin_shipped_editor", hide_index=True, disabled=shipped.columns.drop("선택"), column_order=("선택", "발주번호", "주문일시", "지점명", "건수", "합계금액(원)", "상태"), column_config={"합계금액(원)": st.column_config.NumberColumn()})
        st.session_state.admin_shipped_selection = dict(zip(edited_shipped['발주번호'], edited_shipped['선택']))
        selected_shipped_ids = [k for k, v in st.session_state.admin_shipped_selection.items() if v]

        if st.button("↩️ 접수 상태로 변경", disabled=not selected_shipped_ids, key="admin_revert_btn"):
            if update_order_status(selected_shipped_ids, "접수", st.session_state.auth["name"]):
                st.session_state.success_message = f"{len(selected_shipped_ids)}건이 접수 상태로 변경되었습니다."; st.rerun()
    
    v_spacer(16)
    with st.container(border=True):
        st.markdown("##### 📄 발주 품목 상세 조회")
        total_selected = selected_pending_ids + selected_shipped_ids
        if len(total_selected) == 1:
            target_id = total_selected[0]
            st.markdown(f"**선택된 발주번호:** `{target_id}`")
            target_df = df_all[df_all["발주번호"] == target_id]
            target_status = target_df.iloc[0]["상태"]
            
            df_display = target_df.copy().rename(columns={"단가": "단가(원)", "공급가액": "공급가액(원)", "세액": "세액(원)", "합계금액": "합계금액(원)"})
            display_cols = ["품목코드", "품목명", "단위", "수량", "단가(원)", "공급가액(원)", "세액(원)", "합계금액(원)"]
            st.dataframe(df_display[display_cols], hide_index=True, use_container_width=True, 
                         column_config={
                             "단가(원)": st.column_config.NumberColumn(), "공급가액(원)": st.column_config.NumberColumn(), 
                             "세액(원)": st.column_config.NumberColumn(), "합계금액(원)": st.column_config.NumberColumn()
                         })
            
            if target_status == '출고완료':
                v_spacer(10)
                store_id = target_df.iloc[0]["지점ID"]
                store_info_series = store_info_df[store_info_df["지점ID"] == store_id]
                if not store_info_series.empty:
                    store_info = store_info_series.iloc[0]
                    buf = make_trading_statement_excel(target_df, store_info, master_df)
                    st.download_button(f"'{target_id}' 거래명세서 다운로드", data=buf, file_name=f"거래명세서_{store_info.get('상호명')}_{target_id}.xlsx", mime="application/vnd.ms-excel", use_container_width=True)
        else:
            st.info("상세 내용을 보려면 위 목록에서 발주를 **하나만** 선택하세요.")

def page_admin_documents(store_info_df: pd.DataFrame, master_df: pd.DataFrame):
    st.subheader("📑 증빙서류 다운로드")
    df = load_orders_df()
    df_completed = df[df["상태"] == "출고완료"]
    if df_completed.empty: st.info("'출고완료' 상태의 발주 데이터가 없습니다."); return
    search_mode = st.radio("조회 방식", ["기간으로 조회", "발주번호로 조회"], key="admin_doc_search_mode", horizontal=True)
    dfv = pd.DataFrame(); doc_type = "거래명세서"
    if search_mode == "기간으로 조회":
        c1, c2, c3, c4 = st.columns(4)
        dt_from = c1.date_input("조회 시작일", date.today() - timedelta(days=30), key="admin_doc_from")
        dt_to = c2.date_input("조회 종료일", date.today(), key="admin_doc_to")
        stores = sorted(df_completed["지점명"].dropna().unique().tolist())
        store_sel = c3.selectbox("지점 선택", stores, key="admin_doc_store")
        doc_type = c4.selectbox("문서 종류", ["거래명세서", "세금계산서"], key="admin_doc_type")
        mask = (pd.to_datetime(df_completed["납품요청일"]).dt.date >= dt_from) & (pd.to_datetime(df_completed["납품요청일"]).dt.date <= dt_to) & (df_completed["지점명"] == store_sel)
        dfv = df_completed[mask].copy()
    else:
        c1, c2 = st.columns([1, 1])
        order_ids = sorted(df_completed["발주번호"].dropna().unique().tolist(), reverse=True)
        order_id_sel = c1.selectbox("발주번호 선택", [""] + order_ids, key="admin_doc_order_id")
        doc_type = c2.selectbox("문서 종류", ["거래명세서", "세금계산서"], key="admin_doc_type_by_id")
        if order_id_sel: dfv = df_completed[df_completed["발주번호"] == order_id_sel].copy()
    if dfv.empty: st.warning("해당 조건으로 조회된 데이터가 없습니다."); return
    st.dataframe(dfv, use_container_width=True, hide_index=True)
    if not dfv.empty:
        store_id = dfv.iloc[0]["지점ID"]; store_name = dfv.iloc[0]["지점명"]
        store_info_series = store_info_df[store_info_df["지점ID"] == store_id]
        if not store_info_series.empty:
            store_info = store_info_series.iloc[0]
            if doc_type == "거래명세서":
                buf = make_trading_statement_excel(dfv, store_info, master_df)
            else:
                buf = make_tax_invoice_excel(dfv, store_info, master_df)
            st.download_button(f"'{store_name}' {doc_type} 다운로드", data=buf, file_name=f"{doc_type}_{store_name}_{now_kst_str('%Y%m%d')}.xlsx", mime="application/vnd.ms-excel", use_container_width=True, type="primary")
        else: st.error("지점 정보를 찾을 수 없어 서류를 생성할 수 없습니다.")

def page_admin_items_price(master_df: pd.DataFrame):
    st.subheader("🏷️ 품목 단가 설정")
    st.caption("단가(VAT 제외)를 수정하거나 품목을 추가/삭제한 후 '변경사항 저장' 버튼을 누르세요. 모든 변경 내역은 로그에 기록됩니다.")
    original_df = master_df.copy()
    with st.form("master_edit_form"):
        df_display = master_df.rename(columns={"단가": "단가(원)"})
        edited = st.data_editor(df_display.assign(삭제=False), hide_index=True, num_rows="dynamic", use_container_width=True, column_config={"단가(원)": st.column_config.NumberColumn()})
        if st.form_submit_button("변경사항 저장", type="primary", use_container_width=True):
            edited.rename(columns={"단가(원)": "단가"}, inplace=True)
            edited['삭제'] = edited['삭제'].fillna(False).astype(bool)
            final_df = edited[~edited["삭제"]].drop(columns=["삭제"])
            if write_master_df(final_df, original_df):
                st.session_state.success_message = "상품마스터가 저장되었습니다."; st.rerun()

def page_admin_sales_inquiry(master_df: pd.DataFrame):
    st.subheader("📈 매출 조회")
    df_orders = load_orders_df()
    df_sales_raw = df_orders[df_orders['상태'] == '출고완료'].copy()
    if df_sales_raw.empty: st.info("'출고완료'된 매출 데이터가 없습니다."); return

    c1, c2, c3 = st.columns(3)
    dt_from = c1.date_input("조회 시작일", date.today().replace(day=1), key="admin_sales_from")
    dt_to = c2.date_input("조회 종료일", date.today(), key="admin_sales_to")
    stores = ["(전체 통합)"] + sorted(df_sales_raw["지점명"].dropna().unique().tolist())
    store_sel = c3.selectbox("조회 지점", stores, key="admin_sales_store")

    mask = (pd.to_datetime(df_sales_raw["납품요청일"]).dt.date >= dt_from) & (pd.to_datetime(df_sales_raw["납품요청일"]).dt.date <= dt_to)
    if store_sel != "(전체 통합)": mask &= (df_sales_raw["지점명"] == store_sel)
    df_sales = df_sales_raw[mask].copy()

    if df_sales.empty: st.warning("해당 조건의 매출 데이터가 없습니다."); st.stop()

    total_sales = df_sales["합계금액"].sum(); total_supply = df_sales["공급가액"].sum(); total_tax = df_sales["세액"].sum()
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("총 매출 (VAT 포함)", f"{total_sales:,}원"); m2.metric("공급가액", f"{total_supply:,}원"); m3.metric("부가세액", f"{total_tax:,}원"); m4.metric("총 발주 건수", f"{df_sales['발주번호'].nunique()} 건")
    st.divider()

    sales_tab1, sales_tab2, sales_tab3 = st.tabs(["📊 종합 분석", "📅 일별 상세", "🗓️ 월별 상세"])
    with sales_tab1:
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("##### 🏢 **지점별 매출 순위**")
            store_sales = df_sales.groupby("지점명")["합계금액"].sum().nlargest(10).reset_index()
            store_sales.rename(columns={"합계금액": "매출액(원)"}, inplace=True)
            max_val = int(store_sales['매출액(원)'].max()) if not store_sales.empty else 1
            st.dataframe(store_sales, use_container_width=True, hide_index=True, column_config={"지점명": "지점", "매출액(원)": st.column_config.NumberColumn()})
        with col2:
            st.markdown("##### 🍔 **품목별 판매 순위 (Top 10)**")
            item_sales = df_sales.groupby("품목명")["수량"].sum().nlargest(10).reset_index()
            st.dataframe(item_sales, use_container_width=True, hide_index=True)

    df_sales['일'] = pd.to_datetime(df_sales['납품요청일']).dt.day
    df_sales['월'] = pd.to_datetime(df_sales['납품요청일']).dt.month
    
    with sales_tab2:
        st.markdown("##### 📅 일별 매출 상세")
        daily_pivot = pd.pivot_table(df_sales, values='합계금액', index='일', columns='지점명', aggfunc='sum', fill_value=0)
        if not daily_pivot.empty:
            daily_pivot['총 합계'] = daily_pivot.sum(axis=1)
            st.dataframe(daily_pivot.style.format("{:,.0f}"))
    
    with sales_tab3:
        st.markdown("##### 🗓️ 월별 매출 상세")
        monthly_pivot = pd.pivot_table(df_sales, values='합계금액', index='월', columns='지점명', aggfunc='sum', fill_value=0)
        if not monthly_pivot.empty:
            monthly_pivot['총 합계'] = monthly_pivot.sum(axis=1)
            st.dataframe(monthly_pivot.style.format("{:,.0f}"))
        
    if st.button("매출 정산표 다운로드", use_container_width=True):
        buf = make_sales_summary_excel(daily_pivot, monthly_pivot, f"매출정산표_{dt_from}~{dt_to}")
        st.download_button("다운로드 시작", data=buf, file_name=f"매출정산표_{dt_from}~{dt_to}.xlsx", mime="application/vnd.ms-excel")

# =============================================================================
# 10) 라우팅
# =============================================================================
if __name__ == "__main__":
    if not require_login(): st.stop()
    init_session_state()
    st.title("📦 식자재 발주 시스템")
    display_feedback()
    user = st.session_state.auth
    
    master_df = load_master_df()
    store_info_df = load_store_info_df()
    orders_df = load_orders_df()

    if user["role"] == "admin":
        tabs = st.tabs(["📋 발주요청 조회·수정", "📈 매출 조회", "📑 증빙서류 다운로드", "🏷️ 품목 단가 설정"])
        with tabs[0]: page_admin_unified_management(orders_df, store_info_df, master_df)
        with tabs[1]: page_admin_sales_inquiry(master_df)
        with tabs[2]: page_admin_documents(store_info_df, master_df)
        with tabs[3]: page_admin_items_price(master_df)
    else: # store
        tabs = st.tabs(["🛒 발주 요청", "🧾 발주 조회·수정", "📑 증빙서류 다운로드", "🏷️ 품목 단가 조회"])
        with tabs[0]: page_store_register_confirm(master_df)
        with tabs[1]: page_store_orders_change(store_info_df, master_df)
        with tabs[2]: page_store_documents(store_info_df, master_df)
        with tabs[3]: page_store_master_view(master_df)
